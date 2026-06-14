#!/usr/bin/env python3
"""
PoC — harness de medición de normalizar_bloque (M21 · arco H124, BLUEPRINT §2)
==============================================================================

NO sella nada. Corre la EXTRACCIÓN del parser sobre el bloque bajo 4 configs
(baseline / +headers / +guion / +ambos) sobre la MISMA muestra y mide el efecto.

⚠️  Se corre LOCAL (tu máquina): necesita el corpus `.md` + los índices
    persistidos. El sandbox de Claude NO tiene el corpus (es pesado), por eso
    este harness no se ejecutó allá; allá solo se validó normalizar_bloque y se
    dimensionó el premio (17/300 gold, 100% canal-truncado).

Reconstrucción del bloque (fiel a procesar_archivo): usa los índices YA
refinados que persiste el parser en csjn_casos.csv (`source_file`,
`linea_inicio`, `linea_fin_real`) → `lines[linea_inicio : linea_fin_real+1]`
reproduce el bloque final que ve la extracción. Sobre esa copia corre
normalizar_bloque(**cfg) y luego la misma secuencia que procesar_archivo:
zonificar → dictamen/votos → resolver_dispositivo → classify_outcome.

Métricas (BLUEPRINT §2):
  - por_ello no-vacío  ·  sin_dispositivo count
  - distribución outcome  ·  is_merit (vía _ctx si está)
  - exactitud vs gold (cod_disposicion / cod_es_revision_fondo)  [necesita el HOOK M20]
  - marginal aislado Δ(+headers) Δ(+guion) y la INTERACCIÓN Δ(+ambos) (no asumir aditividad)
  - recuperación del banco n42 (la cláusula jurisdiccional vuelve)

Uso (desde la raíz del repo):
    python scripts/H124/poc_normalizar.py \
        --corpus corpus \
        --casos output/parser/csjn_casos.csv \
        --gold  scripts/H124/in/planilla_M20_codificar.xlsx \
        --n300  scripts/H124/in/M20_clave_parser_n300.csv \
        --n42   scripts/H124/in/B122_banco_truncado_jurisdiccional_n42.csv

HOOK M20: `parser_disposicion` NO se deriva en parser.py — sale de tu script de
derivación M20 (el que produjo el n300). Importalo en derivar_disposicion()
abajo (ver TODO). Sin el hook, el harness mide igual por_ello/outcome/n42 y la
exactitud vs cod_es_revision_fondo por outcome-merit como proxy, pero la
exactitud vs cod_disposicion queda en blanco.
"""
import sys, csv, argparse, random
from pathlib import Path
from collections import Counter, defaultdict

csv.field_size_limit(1 << 24)

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))                      # normalizar_bloque
from normalizar_bloque import normalizar_bloque, CONFIGS

# parser.py vive en scripts/pipeline; lo agregamos al path para importarlo como módulo.
PIPELINE = HERE.parents[1] / "scripts" / "pipeline" if (HERE.parents[1] / "scripts").exists() else None


def _import_parser(pipeline_dir):
    if pipeline_dir is None or not (Path(pipeline_dir) / "parser.py").exists():
        sys.exit("[ABORT] no encuentro scripts/pipeline/parser.py — pasá --pipeline")
    sys.path.insert(0, str(pipeline_dir))
    import parser as P
    return P


# ── HOOK M20: por_ello/considerando → (disposicion, es_revision_fondo, ...) ───
def derivar_disposicion(P, por_ello_text, considerando_text, outcome, is_merit):
    """TODO(Guillermo): importá acá tu derivación M20 (la que produjo el n300).
    Debe devolver un dict con al menos {'disposicion', 'es_revision_fondo'}.
    Ejemplo:
        from derivar_disposicion_m20 import derivar
        return derivar(por_ello_text, considerando_text, outcome, is_merit)
    Mientras no esté cableado, devuelve None y el harness salta la exactitud
    vs cod_disposicion (mide igual el resto)."""
    return None


# ── reconstrucción + extracción bajo una config ──────────────────────────────
def extraer_bajo_config(P, bloque_crudo, cfg_kwargs):
    """Replica la secuencia de procesar_archivo sobre el bloque normalizado.
    Devuelve dict con por_ello_idx, por_ello_text, outcome, considerando_text."""
    bloque = normalizar_bloque(bloque_crudo, **cfg_kwargs)

    _zonas, _ = P.zonificar_bloque(bloque)
    lineas_dictamen = {k for k, z in enumerate(_zonas) if z == "dictamen"}
    lineas_residuo  = {k for k, z in enumerate(_zonas) if z == "residuo_caso_anterior"}
    _ZF = {"apertura", "cuerpo", "dispositivo", "firma", "voto_separado"}
    lineas_excluir  = {k for k, z in enumerate(_zonas) if z not in _ZF}

    _apertura_tipo, apertura_rel = P.detectar_apertura_en_bloque(bloque)
    (_nsv, _ndis, inicio_votos_indiv, _mv) = P.detectar_votos_disidencias(bloque, lineas_excluir)

    por_ello_idx, por_ello_text = P.resolver_dispositivo(
        bloque, apertura_rel, lineas_dictamen, inicio_votos_indiv)

    _no_cons = set(lineas_dictamen) | lineas_residuo
    if inicio_votos_indiv is not None:
        _no_cons |= set(range(inicio_votos_indiv, len(bloque)))
    considerando_text = P.extraer_considerando(bloque, por_ello_idx, _no_cons)

    outcome = P.classify_outcome(por_ello_text, considerando_text)
    return dict(por_ello_idx=por_ello_idx, por_ello_text=por_ello_text,
                considerando_text=considerando_text, outcome=outcome)


# ── carga de insumos ─────────────────────────────────────────────────────────
def cargar_casos_index(casos_csv):
    """{caso_id: (source_file, linea_inicio, linea_fin_real, tipo_entrada, outcome_base)}"""
    idx = {}
    with open(casos_csv, encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            idx[r["caso_id_canonico"]] = dict(
                source_file=r["source_file"],
                linea_inicio=r["linea_inicio"],
                linea_fin_real=r.get("linea_fin_real") or r.get("linea_fin"),
                tipo_entrada=r.get("tipo_entrada", ""),
                outcome_base=r.get("outcome", ""),
                is_merit=r.get("is_merit_decision", ""),
            )
    return idx


def cargar_gold(xlsx):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True)
    ws = wb["M20"] if "M20" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) if h is not None else "" for h in rows[0]]
    ix = {n: i for i, n in enumerate(hdr)}
    out = {}
    for r in rows[1:]:
        cid = r[ix["caso_id_canonico"]] if "caso_id_canonico" in ix else None
        if not cid:
            continue
        def gv(name):
            i = ix.get(name)
            v = r[i] if i is not None and i < len(r) else None
            return (str(v).strip() if v is not None else "")
        out[str(cid)] = dict(
            cod_disposicion=gv("cod_disposicion"),
            cod_es_revision_fondo=gv("cod_es_revision_fondo"),
            cod_reenvia=gv("cod_reenvia"),
        )
    return out


def reconstruir_bloque(P, corpus_dir, source_file, li, lf):
    p = Path(corpus_dir) / source_file
    if not p.exists():
        return None
    lines = p.read_text(encoding="utf-8").split("\n")
    li = int(li) if str(li).strip() not in ("", "None") else 0
    lf = int(lf) if str(lf).strip() not in ("", "None") else len(lines) - 1
    return P.construir_bloque_desde_localizacion(lines, li, lf)


# ── corrida de una muestra bajo las 4 configs ────────────────────────────────
def correr_muestra(P, corpus_dir, casos_idx, ids):
    """Devuelve {config: [registro,...]} con la extracción por caso."""
    res = {cfg: [] for cfg in CONFIGS}
    faltan = 0
    for cid in ids:
        meta = casos_idx.get(cid)
        if not meta or (meta["tipo_entrada"] or "").startswith("sumario"):
            continue
        bloque = reconstruir_bloque(P, corpus_dir, meta["source_file"],
                                    meta["linea_inicio"], meta["linea_fin_real"])
        if not bloque:
            faltan += 1
            continue
        for cfg, kw in CONFIGS.items():
            r = extraer_bajo_config(P, bloque, kw)
            r["caso_id"] = cid
            res[cfg].append(r)
    if faltan:
        print(f"[warn] {faltan} casos sin bloque reconstruible (¿source_file fuera de --corpus?)")
    return res


def resumen_extraccion(registros):
    n = len(registros)
    vacios = sum(1 for r in registros if not r["por_ello_text"].strip())
    outc = Counter(r["outcome"] for r in registros)
    return dict(n=n, por_ello_vacio=vacios, por_ello_ok=n - vacios, outcome=dict(outc))


def exactitud_vs_gold(P, registros, gold):
    """Exactitud vs cod_disposicion / cod_es_revision_fondo (si el HOOK M20 está)."""
    tot_disp = ac_disp = tot_rev = ac_rev = sin_hook = 0
    for r in registros:
        g = gold.get(r["caso_id"])
        if not g:
            continue
        d = derivar_disposicion(P, r["por_ello_text"], r["considerando_text"],
                                r["outcome"], None)
        if d is None:
            sin_hook += 1
            continue
        if g["cod_disposicion"]:
            tot_disp += 1
            ac_disp += int(d.get("disposicion") == g["cod_disposicion"])
        if g["cod_es_revision_fondo"]:
            tot_rev += 1
            ac_rev += int(str(d.get("es_revision_fondo")).lower() == g["cod_es_revision_fondo"].lower())
    return dict(disp=(ac_disp, tot_disp), rev=(ac_rev, tot_rev), sin_hook=sin_hook)


def main():
    ap = argparse.ArgumentParser(description="PoC normalizar_bloque — harness 4 configs (M21/H124)")
    ap.add_argument("--corpus", required=True, help="dir con los .md")
    ap.add_argument("--casos", required=True, help="output/parser/csjn_casos.csv (índices)")
    ap.add_argument("--gold", help="planilla_M20_codificar.xlsx (cod_*)")
    ap.add_argument("--n300", help="M20_clave_parser_n300.csv (define la muestra gold)")
    ap.add_argument("--n42", help="B122_banco_truncado_jurisdiccional_n42.csv")
    ap.add_argument("--pipeline", default=str(PIPELINE) if PIPELINE else None,
                    help="dir de scripts/pipeline (donde está parser.py)")
    ap.add_argument("--corpus-sample", type=int, default=0,
                    help="N casos al azar para deltas corpus-wide (sin gold)")
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    P = _import_parser(args.pipeline)
    casos_idx = cargar_casos_index(args.casos)
    print(f"[ok] parser v{getattr(P,'__version__','?').split()[0]} · {len(casos_idx)} casos en índice\n")

    gold = cargar_gold(args.gold) if args.gold else {}

    # muestra gold = ids del n300
    ids_n300 = []
    if args.n300:
        with open(args.n300, encoding="utf-8") as f:
            ids_n300 = [r["caso_id_canonico"] for r in csv.DictReader(f)]

    # ── corrida n300 (gold) ──────────────────────────────────────────────────
    if ids_n300:
        print(f"══════ MUESTRA GOLD (n300, {len(ids_n300)} casos) ══════")
        res = correr_muestra(P, args.corpus, casos_idx, ids_n300)
        base = resumen_extraccion(res["baseline"])
        print(f"\n{'config':10} {'por_ello_ok':>12} {'sin_disp':>9}   exactitud_vs_gold")
        for cfg in CONFIGS:
            s = resumen_extraccion(res[cfg])
            ex = exactitud_vs_gold(P, res[cfg], gold) if gold else None
            extra = ""
            if ex:
                if ex["sin_hook"]:
                    extra = f"(HOOK M20 sin cablear → {ex['sin_hook']} sin medir)"
                else:
                    ad, td = ex["disp"]; ar, tr = ex["rev"]
                    extra = f"disp {ad}/{td}={ad/td:.3f}  rev {ar}/{tr}={ar/tr:.3f}" if td and tr else ""
            d_ok = s["por_ello_ok"] - base["por_ello_ok"]
            print(f"{cfg:10} {s['por_ello_ok']:>12} {s['por_ello_vacio']:>9}   "
                  f"{extra}   Δpor_ello={d_ok:+d}")
        # interacción
        oks = {cfg: resumen_extraccion(res[cfg])["por_ello_ok"] for cfg in CONFIGS}
        dh = oks["+headers"] - oks["baseline"]
        dg = oks["+guion"] - oks["baseline"]
        dab = oks["+ambos"] - oks["baseline"]
        print(f"\n  marginal headers: {dh:+d} · guion: {dg:+d} · ambos: {dab:+d} · "
              f"aditivo esperado {dh+dg:+d} → {'SUB' if dab<dh+dg else 'SUPER' if dab>dh+dg else 'aditivo'}")

    # ── corrida n42 (recuperación del banco truncado) ────────────────────────
    if args.n42:
        with open(args.n42, encoding="utf-8") as f:
            banco = list(csv.DictReader(f))
        ids42 = [r["caso_id_canonico"] for r in banco]
        base42 = {r["caso_id_canonico"]: r.get("outcome", "") for r in banco}
        print(f"\n══════ BANCO n42 (truncado jurisdiccional, recuperación) ══════")
        res42 = correr_muestra(P, args.corpus, casos_idx, ids42)
        out_by_id = {cfg: {r["caso_id"]: r["outcome"] for r in res42[cfg]} for cfg in CONFIGS}
        for cfg in CONFIGS:
            flips = sum(1 for cid in ids42
                        if cid in out_by_id[cfg] and out_by_id[cfg][cid] != base42.get(cid, ""))
            dist = Counter(out_by_id[cfg].values())
            print(f"  {cfg:10} flips_vs_banco={flips:>3}   outcome_dist={dict(dist)}")

    # ── corpus-wide sample (deltas sin gold) ─────────────────────────────────
    if args.corpus_sample:
        random.seed(args.seed)
        reales = [cid for cid, m in casos_idx.items()
                  if not (m["tipo_entrada"] or "").startswith("sumario")]
        ids = random.sample(reales, min(args.corpus_sample, len(reales)))
        print(f"\n══════ CORPUS-WIDE sample (n={len(ids)}) ══════")
        res = correr_muestra(P, args.corpus, casos_idx, ids)
        base = resumen_extraccion(res["baseline"])
        for cfg in CONFIGS:
            s = resumen_extraccion(res[cfg])
            print(f"  {cfg:10} por_ello_ok={s['por_ello_ok']:>4} "
                  f"sin_disp={s['por_ello_vacio']:>4} Δ={s['por_ello_ok']-base['por_ello_ok']:+d}")

    print("\n[fin] PoC — NO sella nada. Compuerta de decisión: BLUEPRINT §3.")


if __name__ == "__main__":
    main()
